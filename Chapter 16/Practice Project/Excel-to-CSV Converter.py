# Put this py file to the location with the spreadsheets

import openpyxl, csv, os

for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue # skip non-csv files
    else:
        wb = openpyxl.load_workbook(excelFile)
    for sheetName in wb.get_sheet_names():
        # Loop through every sheet in the workbook.
        sheet = wb.get_sheet_by_name(sheetName)
        
        # Create the CSV filename from the Excel filename and sheet title.
        excelFileName = excelFile.rstrip('.xlsx')
        csvFile = open(excelFileName+'-'+sheetName+'.csv','w',newline='')

        # Create the csv.writer object for this CSV file.
        csvWriter=csv.writer(csvFile)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = [] # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                cellValue = sheet.cell(row=rowNum,column=colNum).value
                rowData.append(cellValue)
            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)
        csvFile.close()
        print(excelFileName, sheetName+'.csv saved successfully')
