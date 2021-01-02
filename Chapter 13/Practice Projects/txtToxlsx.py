import openpyxl
from openpyxl.utils import get_column_letter

files = input()
fileList = files.split()

wb = openpyxl.Workbook()
sheet = wb.active

colNum = 1
for file in fileList:
    txtfile = open(file)
    lines = txtfile.readlines()

    rowNum = 1
    
    for line in lines:
        line = line.strip()
        sheet.cell(row = rowNum, column = colNum).value = line
        rowNum += 1

    colLetter = get_column_letter(colNum)
    colNum += 1

wb.save('Text_To_Sheet.xlsx')
print('Saved')
