import openpyxl

wb = openpyxl.load_workbook('grocery.xlsx')

sheet=wb.active

sheetData=[]

for y in range(1,sheet.max_row+1):
    rowData=[]
    for x in range(1,sheet.max_column+1):
        cell = sheet.cell(row=y,column=x).value
        rowData.append(cell)
    sheetData.append(rowData)

wb = openpyxl.Workbook()
sheet = wb.active

colNum=1
for y in sheetData:
    rowNum = 1
    for x in y:
        sheet.cell(row=rowNum,column=colNum).value = x
        rowNum+=1
    colNum+=1

wb.save('inverted.xlsx')
print('Saved')
