import openpyxl

wb = openpyxl.load_workbook('text2sheet.xlsx')
sheet = wb.active

colNum = 1
for column in range(1, sheet.max_column + 1):
    columnData = []    
    for rowNum in range(1, sheet.max_row + 1):
        if sheet.cell(row=rowNum, column=colNum).value:
            columnData.append(sheet.cell(row=rowNum, column=colNum).value)

    file = open('Column'+str(colNum)+'.txt','w')
    for line in columnData:
        file.write(line + '\n')
    file.close()
    print('Saved as '+'Column'+str(colNum)+'.txt')

    colNum += 1
