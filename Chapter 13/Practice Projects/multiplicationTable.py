import openpyxl,sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string
wb = openpyxl.Workbook()
sheet = wb.active

if len(sys.argv) < 2:
    print('Usage: py multiplicationTable.py [number] - copy phrase text')
    sys.exit()
    
number = sys.argv[1]

number=int(number)

fontObj1 = Font(name='Times New Roman', bold=True)

for colNum in range(2, number+2):    
    sheet['A'+str(colNum)].font=fontObj1
    sheet['A'+str(colNum)] = colNum-1

for rowNum in range(2,number+2):
        sheet[get_column_letter(rowNum)+'1'].font=fontObj1
        sheet[get_column_letter(rowNum)+'1']=rowNum-1

for colNum in range(2, number+2):
    for letter in range(2,number+2):
        sheet[get_column_letter(colNum)+str(letter)] = sheet[get_column_letter(colNum)+str('1')].value*sheet[get_column_letter(1)+str(letter)].value

wb.save('mp.xlsx')
print('Ok')

